import os
import re
import pandas as pd
import openpyxl

java_keywords = ['abstract', 'assert', 'boolean', 'break', 'byte', 
                 'case', 'catch', 'char', 'class', 'const', 
                 'continue', 'default', 'do', 'double', 'else',
                 'enum', 'extends', 'final', 'finally', 'float',
                 'for', 'goto', 'if', 'implements', 'import',
                 'instanceof', 'int', 'interface', 'long',
                 'native', 'new', 'package', 'private', 'protected',
                 'public', 'return', 'short', 'static', 'strictfp',
                 'super', 'switch', 'synchronized', 'this', 'throw',
                 'throws', 'transient', 'try', 'void', 'volatile', 'while']

path_to_project = "3 - JRefactory v2.6.24"

package_pattern = re.compile(r"package\s+([\w.]+)\s*;")
class_pattern = re.compile(r"(public|private|protected)?\s*(abstract|static|final)?\s*(class|interface)(\s+(\w+)|\s+(<[\w\s,<>]+>)?\s*\{\s*(public|private|protected)\s+class\s+(\w+)\})\s*(extends\s+([a-zA-Z]+))?\s*(implements\s+([\w\s,<>]+))?")
parent_pattern = re.compile(r"extends\s+([a-zA-Z]+)")
interface_pattern = re.compile(r"implements\s+([\w\s,<>]+)")
method_pattern = re.compile(r"\b(Override)?\s*(((final\s+)?((public|private|protected)\s+)?)|(((public|private|protected)\s+)?(final\s+)?))(synchronized\s+)?(static\s+)?(\w+(\[\])*)\s+(\w+\s?)\s*\(((\s*\w+\s+\w+\s*,?)*)\)\s*\{")
method_constructor_pattern = re.compile(r"\b(public|private|protected)\s*(\w+)\s*\((\w+\s+\w+\s*(,?)\s*)*\)\s*\{")
instantiation_pattern = re.compile(r"=\s*new (\w+)\(((\s*\w+\s*,?)*)\);")
fields_pattern = re.compile(r"\b(public\s+|private\s+|protected\s+|static\s+|final\s+)+\s*(\w+)\s+(\w+)\s*(=\s*([^=\;\n]*))?;")
abstract_pattern = re.compile(r"\b(Override)?\s*(((final\s+)?((public|private|protected)\s+)?)|(((public|private|protected)\s+)?(final\s+)?))(synchronized\s+)?(static\s+)?(abstract\s+)(\w+(\[\])*)\s+(\w+\s?)\s*\(\)\s*;")

project_name = ""
package_name = ""
classes = []

for root, dirs, files in os.walk(path_to_project):
    for filename in files:
        if filename.endswith(".java"):

            with open(os.path.join(root, filename), "r") as f:
                java_code = f.read()
                
                package_match = package_pattern.search(java_code)
                if package_match:
                    package_name = package_match.group(1)

                for class_match in class_pattern.finditer(java_code):
                    
                    visibility_modifier = class_match.group(1) or "default"
                    is_abstract = 1 if class_match.group(2) == "abstract" else 0
                    is_static = 1 if class_match.group(2) == "static" else 0
                    is_final = 1 if class_match.group(2) == "final" else 0
                    class_type = class_match.group(3)
                    class_name = class_match.group(5) or class_match.group(9)
                    is_interface = True if class_type == "interface" else False
                    parent_class = parent_pattern.search(class_match.group(0))
                    parent_name = parent_class.group(1) if parent_class else None
                    interface_classes = interface_pattern.search(class_match.group(0))
                    interface_names = interface_classes.group(1).split(",") if interface_classes else []
                    class_info = {
                        "name": class_name,
                        "type": class_type,
                        "package": package_name,
                        "visibility": visibility_modifier,
                        "is_abstract": is_abstract,
                        "is_static": is_static,
                        "is_final": is_final,
                        "is_interface": is_interface,
                        "parent": parent_name,
                        "interfaces": [name.strip() for name in interface_names],
                        "methods": [],
                        "constructor": [],
                        "instantiations": [],
                        "fields": [],
                        "children": [],
                        "delegations": [],
                        "associated": [],
                        "overrided": [],
                        "aggregation": [],
                        "composition": [],
                        "abstracts": []
                    }
                    classes.append(class_info)
                
                classlistIndex = []
                for class_match in class_pattern.finditer(java_code):
                    classlistIndex.append((class_match.group(5) or class_match.group(9), class_match.span()[0]))
                
                for i in range(len(classlistIndex)):
                    if i < len(classlistIndex) - 1:
                        classTemp = java_code[classlistIndex[i][1]:classlistIndex[i+1][1]]
                    else:
                        classTemp = java_code[classlistIndex[i][1]:]
                    classTempMethod = []
                    classOverlodeMethod = []
                    for c in method_pattern.finditer(classTemp):
                        if "Override" in c.group():
                            method_name = c.group(15)
                            visibility_modifier = c.group(6) or c.group(9) or "default"
                            is_synchronized = True if " synchronized " in c.group() else False
                            is_static = True if " static " in c.group() else False
                            return_type = c.group(13)
                            input_methods = c.group(16)
                            overrided = True if "Override" in c.group() else False
                            final = True if ' final ' in c.group() else False
                            method_info = {
                                "name": method_name,
                                "visibility": visibility_modifier,
                                "is_synchronized": is_synchronized,
                                "is_static": is_static,
                                "return_type": return_type,
                                "input_methods": input_methods,
                                "overrided": overrided,
                                "final":final
                            }
                            classOverlodeMethod.append(method_info)
                        else:
                            method_name = c.group(15)
                            visibility_modifier = c.group(6) or c.group(9) or "default"
                            is_synchronized = True if " synchronized " in c.group() else False
                            is_static = True if " static " in c.group() else False
                            return_type = c.group(13)
                            input_methods = c.group(16)
                            overrided = True if "Override" in c.group() else False
                            final = True if ' final ' in c.group() else False
                            method_info = {
                                "name": method_name,
                                "visibility": visibility_modifier,
                                "is_synchronized": is_synchronized,
                                "is_static": is_static,
                                "return_type": return_type,
                                "input_methods": input_methods,
                                "overrided": overrided,
                                "final":final
                            }
                            classTempMethod.append(method_info)
                    classConstructor = []
                    for c in method_constructor_pattern.finditer(classTemp):
                        method_name = c.group(2)
                        visibility_modifier = c.group(1) or "default"
                        input_methods = c.group(3)
                        method_info = {
                            "name": method_name,
                            "visibility": visibility_modifier
                        }
                        classConstructor.append(method_info)
                    instantiations = []
                    for c in instantiation_pattern.finditer(classTemp):
                        object_name = c.group(1)
                        instantiations.append(object_name)
                    fields = []
                    for c in fields_pattern.finditer(classTemp):
                        field_name = c.group(3)
                        field_type = c.group(2)
                        field_info = {
                            "name": field_name,
                            "type": field_type
                        }
                        fields.append(field_info)
                    abstracts = []
                    for c in abstract_pattern.finditer(classTemp):
                        method_name = c.group(16)
                        abstracts.append(method_name)
                    
                    delegations = {}
                    for match in re.findall(r"\b(\w+)\.(\w+)\(", classTemp):
                        object_name, method_name = match
                        if object_name != "this":
                            if object_name not in delegations:
                                delegations[object_name] = []
                            delegations[object_name].append(method_name)

                    for classinfo in classes:
                        if classinfo["name"] == classlistIndex[i][0]:
                            classinfo["methods"] = classTempMethod
                            classinfo["overrided"] = classOverlodeMethod
                            classinfo["constructor"] = classConstructor
                            classinfo["instantiations"] = instantiations
                            classinfo["fields"] = fields
                            classinfo["abstracts"] = abstracts
                            classinfo["delegations"] = delegations
                            break
                    

for class_info in classes:
    if class_info['parent']:
        parent_name = class_info['parent']
        for child_class in classes:
            if child_class['name'] != parent_name:
                continue
            child_class['children'].append(class_info['name'])

path_to_project = "/Users/amin/Downloads/1 - QuickUML 2001"

java_files = []
for root, dirs, files in os.walk(path_to_project):
    for file in files:
        if file.endswith(".java"):
            java_files.append(os.path.join(root, file))

class_names = set()
imports = {}
for java_file in java_files:
    with open(java_file) as f:
        content = f.read()
        matches = re.findall(r"class\s+(\w+)\s*", content)
        class_names.update(matches)
        matches = re.findall(r"import\s+([\w\.]+);", content)
        for match in matches:
            parts = match.split('.')
            package_name = '.'.join(parts[:-1])
            class_name = parts[-1]
            if package_name not in imports:
                imports[package_name] = [class_name]
            else:
                imports[package_name].append(class_name)

associations = {}
for class_name in class_names:
    associated_classes = []
    for package_name, imported_classes in imports.items():
        if class_name in imported_classes:
            for imported_class in imported_classes:
                if imported_class != class_name and imported_class in class_names:
                    if imported_class in associated_classes:
                        continue
                    associated_classes.append(imported_class)
    associations[class_name] = associated_classes

for class_name in associations:
    for temp in classes:
        if temp["name"] == class_name:
            temp["associated"] = associations[class_name]


# for class_info in classes:
#     print(f"Project name: {os.path.basename(path_to_project)}")
#     print(f"Package name: {package_name}")
#     print(f"\nClass: {class_info['name']}")
#     print(f"Type: {class_info['type']}")
#     print(f"Package: {class_info['package']}")
#     print(f"Visibility: {class_info['visibility']}")
#     print(f"Is abstract: {class_info['is_abstract']}")
#     if class_info['is_static']:
#         print(f"Is static: {class_info['is_static']}")
#     print(f"Is final: {class_info['is_final']}")
#     print(f"Is interface: {class_info['is_interface']}")
#     if class_info['parent']:
#         print(f"Parent: {class_info['parent']}")
#     if class_info['interfaces']:
#         print(f"Implements: {', '.join(class_info['interfaces'])}")
#     if class_info['methods']:
#         print("Methods:")
#         for method in class_info['methods']:
#             print(f"\tname: {method['name']}")
#             for method_info in method:
#                 if method[method_info] and method_info != 'name':
#                     print(f"\t\t{method_info}: {method[method_info]}")
#     if class_info['constructor']:
#         print("constructor:")
#         for method in class_info['constructor']:
#             print(f"\tname: {method['name']}")
#             for method_info in method:
#                 if method[method_info] and method_info != 'name':
#                     print(f"\t\t{method_info}: {method[method_info]}")
#     if class_info['children']:
#         print(f"Children: {', '.join(class_info['children'])}")
#     if class_info['instantiations']:
#         print('instantiations:')
#         for instant in class_info['instantiations']:
#             print(f'\t{instant}')
#     if class_info['fields']:
#         print('fields:')
#         for field in class_info['fields']:
#             print(f"\tname: {field['name']}, type: {field['type']}")
#     if class_info['abstracts']:
#         print('abstracts:')
#         for field in class_info['abstracts']:
#             print(f"\tname: {field}")
#     if class_info['delegations']:
#         print('delegations:')
#         for field in class_info['delegations']:
#             print(f"\tname: {field}")
#     if class_info["associated"]:
#         print(f"associated with:")
#         for asso in class_info["associated"]:
#             print("\t",asso)

wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=1, value='Project Name')
ws.cell(row=1, column=2, value='Package_Name')
ws.cell(row=1, column=3, value='Class_Name')
ws.cell(row=1, column=4, value='Class_Type')
ws.cell(row=1, column=5, value='Class_Visibility')
ws.cell(row=1, column=6, value='Class_is_Abstract')
ws.cell(row=1, column=7, value='Class_is_Static')
ws.cell(row=1, column=8, value='Class_is_Final')
ws.cell(row=1, column=9, value='Class_Is_Interface')
ws.cell(row=1, column=10, value='extends')
ws.cell(row=1, column=11, value='implements')
ws.cell(row=1, column=12, value='children')
ws.cell(row=1, column=13, value='constructor')
ws.cell(row=1, column=14, value='Fields')
ws.cell(row=1, column=15, value='Methods')
ws.cell(row=1, column=16, value='override')
ws.cell(row=1, column=17, value='has_static_method')
ws.cell(row=1, column=18, value='has_final_method')
ws.cell(row=1, column=19, value='has_abstract_method')
ws.cell(row=1, column=20, value='Association')
ws.cell(row=1, column=21, value='Aggregation')
ws.cell(row=1, column=22, value='Delegation')
ws.cell(row=1, column=23, value='Composition')
ws.cell(row=1, column=24, value='Instantiation')

i = 0
for class_info in classes:
    if class_info['name'] in java_keywords:
        continue
    row_num = i + 2
    ws.cell(row=row_num, column=1, value=os.path.basename(path_to_project))
    ws.cell(row=row_num, column=2, value=class_info['package'])
    if class_info['name']:
        ws.cell(row=row_num, column=3, value=class_info['name'])
    if class_info['type']:
        ws.cell(row=row_num, column=4, value= 1 if class_info['type'] == 'class' else (2 if class_info['type'] == 'interface' else 3))
    if class_info['visibility']:
        ws.cell(row=row_num, column=5, value= 1 if class_info['visibility'] == 'public' else (2 if class_info['visibility'] == 'private' else 3))
    ws.cell(row=row_num, column=6, value=class_info['is_abstract'])
    ws.cell(row=row_num, column=7, value=class_info['is_static'])
    ws.cell(row=row_num, column=8, value=class_info['is_final'])
    ws.cell(row=row_num, column=9, value=class_info['is_interface'])
    if class_info['parent']:
        ws.cell(row=row_num, column=10, value=class_info['parent'])
    else:
        ws.cell(row=row_num, column=10, value=0)
    if class_info['interfaces']:
        ws.cell(row=row_num, column=11, value=', '.join(class_info['interfaces']))
    else:
        ws.cell(row=row_num, column=11, value=0)
    if class_info['children']:
        ws.cell(row=row_num, column=12, value=', '.join(class_info['children']))
    else:
        ws.cell(row=row_num, column=12, value=0)
    if class_info['constructor']:
        ws.cell(row=row_num, column=13, value=str(class_info['constructor']))
    if class_info['fields']:
        ws.cell(row=row_num, column=14, value=', '.join([f"{field['name']}, type: {field['type']}" for field in class_info['fields']]))
    if class_info['methods']:
        ws.cell(row=row_num, column=15, value=str(class_info['methods']))
    if class_info['overrided']:
        ws.cell(row=row_num, column=16, value=str(class_info['overrided']))
    if class_info['methods']:
        ws.cell(row=row_num, column=17, value=any([method if method['is_static'] else '' for method in class_info['methods']]))
    if class_info['methods']:
        ws.cell(row=row_num, column=18, value=any([method if method['final'] else '' for method in class_info['methods']]))
    if class_info['abstracts']:
        ws.cell(row=row_num, column=19, value=str(class_info['abstracts']))
    if class_info['associated']:
        ws.cell(row=row_num, column=20, value=', '.join(class_info['associated']))
    if class_info['aggregation']:
        ws.cell(row=row_num, column=21, value=', '.join(class_info["aggregation"]))
    if class_info['delegations']:
        ws.cell(row=row_num, column=22, value=', '.join(class_info['delegations']))
    if class_info['composition']:
        ws.cell(row=row_num, column=23, value=', '.join(class_info['composition']))
    if class_info['instantiations']:
        ws.cell(row=row_num, column=24, value=', '.join(class_info['instantiations']))
    i+=1

wb.save(os.path.basename(path_to_project) + '_class_info.xlsx')
